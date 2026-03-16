from fastapi import FastAPI, APIRouter, HTTPException, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import pymysql
from pymysql.cursors import DictCursor
import bcrypt
import jwt

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'madoverai-secret-key-2024-secure')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# MongoDB connection (fallback)
mongo_url = os.environ['MONGO_URL']
mongo_client = AsyncIOMotorClient(mongo_url)
mongo_db = mongo_client[os.environ['DB_NAME']]

# MySQL connection configuration
MYSQL_CONFIG = {
    'host': os.environ.get('MYSQL_HOST', 'madoverai.cdam6io6a2o3.eu-north-1.rds.amazonaws.com'),
    'port': int(os.environ.get('MYSQL_PORT', 13306)),
    'user': os.environ.get('MYSQL_USER', 'admin'),
    'password': os.environ.get('MYSQL_PASSWORD', 'm94IHMwmhb1SHItnl3zP'),
    'database': os.environ.get('MYSQL_DATABASE', 'madoverai'),
    'charset': 'utf8mb4',
    'cursorclass': DictCursor,
    'connect_timeout': 2,
    'read_timeout': 10,
    'write_timeout': 10,
    'autocommit': True
}

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MySQL availability cache to avoid repeated connection attempts
_mysql_available = None
_mysql_check_time = None
MYSQL_RETRY_INTERVAL = 60  # Retry MySQL check every 60 seconds

def try_mysql_connection():
    """Try to get MySQL connection with caching to avoid repeated timeouts"""
    global _mysql_available, _mysql_check_time
    
    current_time = datetime.now(timezone.utc)
    
    # Skip MySQL if we know it's down and haven't waited long enough
    if _mysql_available is False and _mysql_check_time:
        elapsed = (current_time - _mysql_check_time).total_seconds()
        if elapsed < MYSQL_RETRY_INTERVAL:
            return None
    
    try:
        conn = pymysql.connect(**MYSQL_CONFIG)
        _mysql_available = True
        _mysql_check_time = current_time
        return conn
    except pymysql.Error as e:
        _mysql_available = False
        _mysql_check_time = current_time
        logger.warning(f"MySQL connection failed: {e}")
        return None

def init_mysql_tables(connection):
    """Initialize all MySQL tables"""
    try:
        with connection.cursor() as cursor:
            # Admins table (BrandsXAI employees)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_admins (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(100) NOT NULL UNIQUE,
                    email VARCHAR(255),
                    password_hash VARCHAR(255) NOT NULL,
                    is_active BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_admin_username (username)
                )
            """)
            
            # Brands table (Customers)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_brands (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Features table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_features (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(100) NOT NULL UNIQUE,
                    icon VARCHAR(50),
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Feature Pages table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_feature_pages (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    feature_id INT NOT NULL,
                    name VARCHAR(100) NOT NULL,
                    icon VARCHAR(50),
                    route VARCHAR(100) NOT NULL,
                    display_order INT DEFAULT 0,
                    FOREIGN KEY (feature_id) REFERENCES brandsxai_features(id) ON DELETE CASCADE
                )
            """)
            
            # Users table (Brand users)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(100) NOT NULL UNIQUE,
                    email VARCHAR(255),
                    password_hash VARCHAR(255) NOT NULL,
                    brand_id INT,
                    is_active BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_login TIMESTAMP NULL,
                    FOREIGN KEY (brand_id) REFERENCES brandsxai_brands(id) ON DELETE SET NULL,
                    INDEX idx_user_username (username),
                    INDEX idx_user_brand (brand_id)
                )
            """)
            
            # User Features mapping
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_user_features (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    feature_id INT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES brandsxai_users(id) ON DELETE CASCADE,
                    FOREIGN KEY (feature_id) REFERENCES brandsxai_features(id) ON DELETE CASCADE,
                    UNIQUE KEY unique_user_feature (user_id, feature_id)
                )
            """)
            
            # Campaigns table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_campaigns (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    brand_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    description TEXT,
                    start_date DATE,
                    end_date DATE,
                    target_audience VARCHAR(255),
                    call_script TEXT,
                    status ENUM('draft', 'active', 'paused', 'completed') DEFAULT 'draft',
                    created_by INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (brand_id) REFERENCES brandsxai_brands(id) ON DELETE CASCADE,
                    INDEX idx_campaign_brand (brand_id),
                    INDEX idx_campaign_status (status)
                )
            """)
            
            # Opportunities/Leads table for campaigns
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_opportunities (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    campaign_id INT NOT NULL,
                    brand_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    phone VARCHAR(50),
                    email VARCHAR(255),
                    business_name VARCHAR(255),
                    opportunity_value DECIMAL(10, 2) DEFAULT 0.00,
                    stage ENUM('dialing', 'interested', 'not_interested', 'callback', 'store_visit', 'invalid_number') DEFAULT 'dialing',
                    notes TEXT,
                    last_called_at TIMESTAMP NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (campaign_id) REFERENCES brandsxai_campaigns(id) ON DELETE CASCADE,
                    FOREIGN KEY (brand_id) REFERENCES brandsxai_brands(id) ON DELETE CASCADE,
                    INDEX idx_opp_campaign (campaign_id),
                    INDEX idx_opp_stage (stage)
                )
            """)
            
            # Leads table (for contact form)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_leads (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    email VARCHAR(255) NOT NULL,
                    company VARCHAR(255),
                    message TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_lead_email (email)
                )
            """)
            
            # Claim Processing Sessions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_claim_sessions (
                    id VARCHAR(36) PRIMARY KEY,
                    user_id INT NOT NULL,
                    brand_id INT NOT NULL,
                    title VARCHAR(255) DEFAULT 'New Claim Session',
                    status ENUM('active', 'completed', 'archived') DEFAULT 'active',
                    extracted_codes JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_claim_user (user_id),
                    INDEX idx_claim_brand (brand_id),
                    INDEX idx_claim_status (status)
                )
            """)
            
            # Claim Processing Messages table (chat history)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS brandsxai_claim_messages (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    session_id VARCHAR(36) NOT NULL,
                    role ENUM('user', 'assistant', 'system') NOT NULL,
                    content TEXT NOT NULL,
                    file_info JSON,
                    codes_extracted JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_msg_session (session_id)
                )
            """)
            
            # Insert default admin if not exists
            cursor.execute("SELECT id FROM brandsxai_admins WHERE username = 'madoveradmin'")
            if not cursor.fetchone():
                password_hash = bcrypt.hashpw('admin@123'.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
                cursor.execute(
                    "INSERT INTO brandsxai_admins (username, email, password_hash) VALUES (%s, %s, %s)",
                    ('madoveradmin', 'admin@madoverai.com', password_hash)
                )
                logger.info("Created default BrandsXAI admin")
            
            # Insert default features if not exists
            cursor.execute("SELECT id FROM brandsxai_features WHERE name = 'Voice AI'")
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO brandsxai_features (name, icon, description) VALUES (%s, %s, %s)",
                    ('Voice AI', 'Phone', 'Voice AI solutions for customer interactions')
                )
                voice_ai_id = cursor.lastrowid
                
                # Insert Voice AI pages
                cursor.execute(
                    "INSERT INTO brandsxai_feature_pages (feature_id, name, icon, route, display_order) VALUES (%s, %s, %s, %s, %s)",
                    (voice_ai_id, 'Contacts', 'Users', '/dashboard/voice-ai/contacts', 1)
                )
                cursor.execute(
                    "INSERT INTO brandsxai_feature_pages (feature_id, name, icon, route, display_order) VALUES (%s, %s, %s, %s, %s)",
                    (voice_ai_id, 'Dashboards', 'LayoutDashboard', '/dashboard/voice-ai/dashboards', 2)
                )
                cursor.execute(
                    "INSERT INTO brandsxai_feature_pages (feature_id, name, icon, route, display_order) VALUES (%s, %s, %s, %s, %s)",
                    (voice_ai_id, 'Session', 'Clock', '/dashboard/voice-ai/session', 3)
                )
                cursor.execute(
                    "INSERT INTO brandsxai_feature_pages (feature_id, name, icon, route, display_order) VALUES (%s, %s, %s, %s, %s)",
                    (voice_ai_id, 'Campaign', 'Megaphone', '/dashboard/voice-ai/campaign', 4)
                )
                logger.info("Created Voice AI feature with pages")
            
            cursor.execute("SELECT id FROM brandsxai_features WHERE name = 'Claim Processing'")
            claim_feature = cursor.fetchone()
            if not claim_feature:
                cursor.execute(
                    "INSERT INTO brandsxai_features (name, icon, description) VALUES (%s, %s, %s)",
                    ('Claim Processing', 'FileCheck', 'AI-powered medical claim processing and ICD-10 code extraction')
                )
                claim_id = cursor.lastrowid
                cursor.execute(
                    "INSERT INTO brandsxai_feature_pages (feature_id, name, icon, route, display_order) VALUES (%s, %s, %s, %s, %s)",
                    (claim_id, 'Code Extractor', 'FileSearch', '/dashboard/claim-processing/extractor', 1)
                )
                logger.info("Created Claim Processing feature with pages")
            
            # Insert sample brands if not exists
            cursor.execute("SELECT id FROM brandsxai_brands WHERE name = 'Brand X'")
            if not cursor.fetchone():
                cursor.execute("INSERT INTO brandsxai_brands (name) VALUES (%s)", ('Brand X',))
                cursor.execute("INSERT INTO brandsxai_brands (name) VALUES (%s)", ('Brand Y',))
                logger.info("Created sample brands")
            
            connection.commit()
            logger.info("MySQL tables initialized successfully")
            return True
    except pymysql.Error as e:
        logger.error(f"MySQL table initialization error: {e}")
        return False

async def init_mongodb_collections():
    """Initialize MongoDB collections with default data"""
    try:
        # Create indexes
        await mongo_db.brandsxai_admins.create_index("username", unique=True)
        await mongo_db.brandsxai_users.create_index("username", unique=True)
        await mongo_db.brandsxai_brands.create_index("name", unique=True)
        await mongo_db.brandsxai_features.create_index("name", unique=True)
        
        # Default admin
        admin = await mongo_db.brandsxai_admins.find_one({"username": "madoveradmin"})
        if not admin:
            password_hash = bcrypt.hashpw('admin@123'.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
            await mongo_db.brandsxai_admins.insert_one({
                "id": 1, "username": "madoveradmin", "email": "admin@madoverai.com",
                "password_hash": password_hash, "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        # Default features
        voice_ai = await mongo_db.brandsxai_features.find_one({"name": "Voice AI"})
        if not voice_ai:
            await mongo_db.brandsxai_features.insert_one({
                "id": 1, "name": "Voice AI", "icon": "Phone", "description": "Voice AI solutions",
                "pages": [
                    {"id": 1, "name": "Contacts", "icon": "Users", "route": "/dashboard/voice-ai/contacts", "display_order": 1},
                    {"id": 2, "name": "Dashboards", "icon": "LayoutDashboard", "route": "/dashboard/voice-ai/dashboards", "display_order": 2},
                    {"id": 3, "name": "Session", "icon": "Clock", "route": "/dashboard/voice-ai/session", "display_order": 3},
                    {"id": 4, "name": "Campaign", "icon": "Megaphone", "route": "/dashboard/voice-ai/campaign", "display_order": 4}
                ]
            })
        else:
            # Update existing Voice AI to add Campaign page if missing
            if not any(p.get('name') == 'Campaign' for p in voice_ai.get('pages', [])):
                pages = voice_ai.get('pages', [])
                pages.append({"id": 4, "name": "Campaign", "icon": "Megaphone", "route": "/dashboard/voice-ai/campaign", "display_order": 4})
                await mongo_db.brandsxai_features.update_one(
                    {"name": "Voice AI"},
                    {"$set": {"pages": pages}}
                )
        
        claim = await mongo_db.brandsxai_features.find_one({"name": "Claim Processing"})
        if not claim:
            await mongo_db.brandsxai_features.insert_one({
                "id": 2, "name": "Claim Processing", "icon": "FileCheck", 
                "description": "AI-powered medical claim processing and ICD-10 code extraction",
                "pages": [
                    {"id": 1, "name": "Code Extractor", "icon": "FileSearch", "route": "/dashboard/claim-processing/extractor", "display_order": 1}
                ]
            })
        else:
            # Update existing Claim Processing to add pages if missing
            if not claim.get('pages'):
                await mongo_db.brandsxai_features.update_one(
                    {"name": "Claim Processing"},
                    {"$set": {"pages": [
                        {"id": 1, "name": "Code Extractor", "icon": "FileSearch", "route": "/dashboard/claim-processing/extractor", "display_order": 1}
                    ]}}
                )
        
        # Create claim processing indexes
        await mongo_db.brandsxai_claim_sessions.create_index("user_id")
        await mongo_db.brandsxai_claim_sessions.create_index("brand_id")
        await mongo_db.brandsxai_claim_messages.create_index("session_id")
        
        # Default brands
        brand_x = await mongo_db.brandsxai_brands.find_one({"name": "Brand X"})
        if not brand_x:
            await mongo_db.brandsxai_brands.insert_one({"id": 1, "name": "Brand X", "created_at": datetime.now(timezone.utc).isoformat()})
            await mongo_db.brandsxai_brands.insert_one({"id": 2, "name": "Brand Y", "created_at": datetime.now(timezone.utc).isoformat()})
        
        logger.info("MongoDB collections initialized")
    except Exception as e:
        logger.error(f"MongoDB initialization error: {e}")

# Create the main app
app = FastAPI(title="BrandsXAI API", version="1.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

# ==================== PYDANTIC MODELS ====================

class LoginRequest(BaseModel):
    username: str
    password: str

class AdminLoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    admin: dict
    is_admin: bool = True

class UserLoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict
    brand: dict
    features: List[dict]
    is_admin: bool = False

class CreateUserRequest(BaseModel):
    username: str
    password: str
    email: Optional[str] = None
    brand_id: int
    feature_ids: List[int]

class UpdateUserFeaturesRequest(BaseModel):
    feature_ids: List[int]

class CreateBrandRequest(BaseModel):
    name: str

class LeadCreate(BaseModel):
    name: str
    email: EmailStr
    company: Optional[str] = None
    message: Optional[str] = None

# ==================== CLAIM PROCESSING MODELS ====================

class ClaimSessionCreate(BaseModel):
    title: Optional[str] = "New Claim Session"

class ClaimMessage(BaseModel):
    content: str
    file_data: Optional[List[dict]] = None  # [{filename, content_type, base64_data}]

class ICD10Code(BaseModel):
    code: str
    description: str
    source_text: Optional[str] = None
    confidence: Optional[float] = None

class ClaimSessionResponse(BaseModel):
    id: str
    title: str
    status: str
    extracted_codes: List[dict]
    created_at: str
    updated_at: str

class CodeUpdateRequest(BaseModel):
    codes: List[dict]  # [{code, description}]

# ==================== JWT HELPERS ====================

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(token: str) -> Optional[dict]:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except:
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Optional[dict]:
    if not credentials:
        return None
    return verify_token(credentials.credentials)

# ==================== ADMIN AUTH ====================

@api_router.post("/admin/login")
async def admin_login(request: LoginRequest):
    """BrandsXAI Admin login"""
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_admins WHERE username = %s AND is_active = TRUE", (request.username,))
                admin = cursor.fetchone()
                if admin and bcrypt.checkpw(request.password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
                    token = create_access_token({"sub": str(admin['id']), "username": admin['username'], "type": "admin"})
                    return {"access_token": token, "token_type": "bearer", "admin": {"id": admin['id'], "username": admin['username']}, "is_admin": True, "db_source": "mysql"}
        except Exception as e:
            logger.warning(f"MySQL admin login error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    admin = await mongo_db.brandsxai_admins.find_one({"username": request.username, "is_active": True})
    if admin and bcrypt.checkpw(request.password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
        token = create_access_token({"sub": str(admin.get('id', 1)), "username": admin['username'], "type": "admin"})
        return {"access_token": token, "token_type": "bearer", "admin": {"id": admin.get('id', 1), "username": admin['username']}, "is_admin": True, "db_source": "mongodb"}
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

# ==================== USER AUTH ====================

@api_router.post("/auth/login")
async def user_login(request: LoginRequest):
    """Brand user login"""
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                # Get user with brand
                cursor.execute("""
                    SELECT u.*, b.name as brand_name 
                    FROM brandsxai_users u 
                    LEFT JOIN brandsxai_brands b ON u.brand_id = b.id 
                    WHERE u.username = %s AND u.is_active = TRUE
                """, (request.username,))
                user = cursor.fetchone()
                
                if user and bcrypt.checkpw(request.password.encode('utf-8'), user['password_hash'].encode('utf-8')):
                    # Get user features with pages
                    cursor.execute("""
                        SELECT f.*, GROUP_CONCAT(
                            CONCAT(fp.id, ':', fp.name, ':', fp.icon, ':', fp.route, ':', fp.display_order) 
                            ORDER BY fp.display_order
                        ) as pages
                        FROM brandsxai_user_features uf
                        JOIN brandsxai_features f ON uf.feature_id = f.id
                        LEFT JOIN brandsxai_feature_pages fp ON f.id = fp.feature_id
                        WHERE uf.user_id = %s
                        GROUP BY f.id
                    """, (user['id'],))
                    features_raw = cursor.fetchall()
                    
                    features = []
                    for f in features_raw:
                        feature = {"id": f['id'], "name": f['name'], "icon": f['icon'], "description": f['description'], "pages": []}
                        if f['pages']:
                            for p in f['pages'].split(','):
                                parts = p.split(':')
                                if len(parts) >= 5:
                                    feature['pages'].append({"id": int(parts[0]), "name": parts[1], "icon": parts[2], "route": parts[3], "display_order": int(parts[4])})
                        features.append(feature)
                    
                    # Update last login
                    cursor.execute("UPDATE brandsxai_users SET last_login = NOW() WHERE id = %s", (user['id'],))
                    mysql_conn.commit()
                    
                    token = create_access_token({
                        "sub": str(user['id']), "username": user['username'], "type": "user",
                        "brand_id": user['brand_id'], "brand_name": user['brand_name']
                    })
                    
                    return {
                        "access_token": token, "token_type": "bearer",
                        "user": {"id": user['id'], "username": user['username'], "email": user['email']},
                        "brand": {"id": user['brand_id'], "name": user['brand_name']},
                        "features": features, "is_admin": False, "db_source": "mysql"
                    }
        except Exception as e:
            logger.warning(f"MySQL user login error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    user = await mongo_db.brandsxai_users.find_one({"username": request.username, "is_active": True})
    if user and bcrypt.checkpw(request.password.encode('utf-8'), user['password_hash'].encode('utf-8')):
        brand = await mongo_db.brandsxai_brands.find_one({"id": user.get('brand_id')})
        
        # Get features
        feature_ids = user.get('feature_ids', [])
        features = []
        async for f in mongo_db.brandsxai_features.find({"id": {"$in": feature_ids}}):
            features.append({"id": f['id'], "name": f['name'], "icon": f['icon'], "description": f.get('description', ''), "pages": f.get('pages', [])})
        
        await mongo_db.brandsxai_users.update_one({"_id": user['_id']}, {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}})
        
        token = create_access_token({
            "sub": str(user.get('id', 1)), "username": user['username'], "type": "user",
            "brand_id": user.get('brand_id'), "brand_name": brand['name'] if brand else None
        })
        
        return {
            "access_token": token, "token_type": "bearer",
            "user": {"id": user.get('id', 1), "username": user['username'], "email": user.get('email', '')},
            "brand": {"id": brand['id'], "name": brand['name']} if brand else None,
            "features": features, "is_admin": False, "db_source": "mongodb"
        }
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

# ==================== ADMIN: USER MANAGEMENT ====================

@api_router.post("/admin/users")
async def create_user(request: CreateUserRequest, current_user: dict = Depends(get_current_user)):
    """Create a new brand user (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    password_hash = bcrypt.hashpw(request.password.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                # Create user
                cursor.execute(
                    "INSERT INTO brandsxai_users (username, email, password_hash, brand_id) VALUES (%s, %s, %s, %s)",
                    (request.username, request.email, password_hash, request.brand_id)
                )
                user_id = cursor.lastrowid
                
                # Add feature access
                for feature_id in request.feature_ids:
                    cursor.execute(
                        "INSERT INTO brandsxai_user_features (user_id, feature_id) VALUES (%s, %s)",
                        (user_id, feature_id)
                    )
                
                mysql_conn.commit()
                return {"message": "User created", "user_id": user_id, "db_source": "mysql"}
        except pymysql.IntegrityError:
            raise HTTPException(status_code=400, detail="Username already exists")
        except Exception as e:
            logger.error(f"MySQL create user error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    existing = await mongo_db.brandsxai_users.find_one({"username": request.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    last_user = await mongo_db.brandsxai_users.find_one(sort=[("id", -1)])
    new_id = (last_user.get('id', 0) + 1) if last_user else 1
    
    await mongo_db.brandsxai_users.insert_one({
        "id": new_id, "username": request.username, "email": request.email,
        "password_hash": password_hash, "brand_id": request.brand_id,
        "feature_ids": request.feature_ids, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "User created", "user_id": new_id, "db_source": "mongodb"}

@api_router.get("/admin/users")
async def get_all_users(current_user: dict = Depends(get_current_user)):
    """Get all users (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT u.id, u.username, u.email, u.is_active, u.created_at, u.last_login,
                           b.id as brand_id, b.name as brand_name,
                           GROUP_CONCAT(f.id) as feature_ids, GROUP_CONCAT(f.name) as feature_names
                    FROM brandsxai_users u
                    LEFT JOIN brandsxai_brands b ON u.brand_id = b.id
                    LEFT JOIN brandsxai_user_features uf ON u.id = uf.user_id
                    LEFT JOIN brandsxai_features f ON uf.feature_id = f.id
                    GROUP BY u.id
                    ORDER BY u.created_at DESC
                """)
                users = cursor.fetchall()
                
                result = []
                for u in users:
                    result.append({
                        "id": u['id'], "username": u['username'], "email": u['email'],
                        "is_active": u['is_active'], "created_at": u['created_at'], "last_login": u['last_login'],
                        "brand": {"id": u['brand_id'], "name": u['brand_name']} if u['brand_id'] else None,
                        "features": [{"id": int(fid), "name": fn} for fid, fn in zip(
                            (u['feature_ids'] or '').split(','), (u['feature_names'] or '').split(',')
                        ) if fid and fn]
                    })
                
                return {"users": result, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get users error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    users = []
    async for u in mongo_db.brandsxai_users.find({}, {"_id": 0, "password_hash": 0}):
        brand = None
        if u.get('brand_id'):
            brand = await mongo_db.brandsxai_brands.find_one({"id": u.get('brand_id')}, {"_id": 0})
        features = []
        for fid in u.get('feature_ids', []):
            f = await mongo_db.brandsxai_features.find_one({"id": fid}, {"_id": 0})
            if f:
                features.append({"id": f['id'], "name": f['name']})
        users.append({
            "id": u.get('id'),
            "username": u.get('username'),
            "email": u.get('email'),
            "is_active": u.get('is_active', True),
            "created_at": u.get('created_at'),
            "last_login": u.get('last_login'),
            "brand": {"id": brand['id'], "name": brand['name']} if brand else None,
            "features": features
        })
    
    logger.info(f"MongoDB: Found {len(users)} users")
    return {"users": users, "db_source": "mongodb"}

@api_router.put("/admin/users/{user_id}/features")
async def update_user_features(user_id: int, request: UpdateUserFeaturesRequest, current_user: dict = Depends(get_current_user)):
    """Update user's feature access (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                # Remove existing features
                cursor.execute("DELETE FROM brandsxai_user_features WHERE user_id = %s", (user_id,))
                
                # Add new features
                for feature_id in request.feature_ids:
                    cursor.execute(
                        "INSERT INTO brandsxai_user_features (user_id, feature_id) VALUES (%s, %s)",
                        (user_id, feature_id)
                    )
                
                mysql_conn.commit()
                return {"message": "Features updated", "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL update features error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    await mongo_db.brandsxai_users.update_one(
        {"id": user_id},
        {"$set": {"feature_ids": request.feature_ids}}
    )
    return {"message": "Features updated", "db_source": "mongodb"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: int, current_user: dict = Depends(get_current_user)):
    """Delete/deactivate a user (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("UPDATE brandsxai_users SET is_active = FALSE WHERE id = %s", (user_id,))
                mysql_conn.commit()
                return {"message": "User deactivated", "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL delete user error: {e}")
        finally:
            mysql_conn.close()
    
    await mongo_db.brandsxai_users.update_one({"id": user_id}, {"$set": {"is_active": False}})
    return {"message": "User deactivated", "db_source": "mongodb"}

# ==================== ADMIN: BRAND MANAGEMENT ====================

@api_router.get("/admin/brands")
async def get_brands(current_user: dict = Depends(get_current_user)):
    """Get all brands (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_brands ORDER BY name")
                return {"brands": cursor.fetchall(), "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get brands error: {e}")
        finally:
            mysql_conn.close()
    
    brands = await mongo_db.brandsxai_brands.find({}, {"_id": 0}).to_list(100)
    return {"brands": brands, "db_source": "mongodb"}

@api_router.post("/admin/brands")
async def create_brand(request: CreateBrandRequest, current_user: dict = Depends(get_current_user)):
    """Create a new brand (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("INSERT INTO brandsxai_brands (name) VALUES (%s)", (request.name,))
                mysql_conn.commit()
                return {"message": "Brand created", "brand_id": cursor.lastrowid, "db_source": "mysql"}
        except pymysql.IntegrityError:
            raise HTTPException(status_code=400, detail="Brand already exists")
        except Exception as e:
            logger.error(f"MySQL create brand error: {e}")
        finally:
            mysql_conn.close()
    
    existing = await mongo_db.brandsxai_brands.find_one({"name": request.name})
    if existing:
        raise HTTPException(status_code=400, detail="Brand already exists")
    
    last_brand = await mongo_db.brandsxai_brands.find_one(sort=[("id", -1)])
    new_id = (last_brand.get('id', 0) + 1) if last_brand else 1
    
    await mongo_db.brandsxai_brands.insert_one({
        "id": new_id, "name": request.name, "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "Brand created", "brand_id": new_id, "db_source": "mongodb"}

# ==================== ADMIN: FEATURE MANAGEMENT ====================

@api_router.get("/admin/features")
async def get_features(current_user: dict = Depends(get_current_user)):
    """Get all features (Admin only)"""
    if not current_user or current_user.get('type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_features ORDER BY name")
                features = cursor.fetchall()
                
                for f in features:
                    cursor.execute("SELECT * FROM brandsxai_feature_pages WHERE feature_id = %s ORDER BY display_order", (f['id'],))
                    f['pages'] = cursor.fetchall()
                
                return {"features": features, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get features error: {e}")
        finally:
            mysql_conn.close()
    
    features = await mongo_db.brandsxai_features.find({}, {"_id": 0}).to_list(100)
    return {"features": features, "db_source": "mongodb"}

# ==================== LEADS ====================

@api_router.post("/leads")
async def create_lead(lead: LeadCreate):
    """Create lead from contact form"""
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO brandsxai_leads (name, email, company, message) VALUES (%s, %s, %s, %s)",
                    (lead.name, lead.email, lead.company, lead.message)
                )
                mysql_conn.commit()
                return {"message": "Lead created", "id": cursor.lastrowid, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL create lead error: {e}")
        finally:
            mysql_conn.close()
    
    last_lead = await mongo_db.brandsxai_leads.find_one(sort=[("id", -1)])
    new_id = (last_lead.get('id', 0) + 1) if last_lead else 1
    
    await mongo_db.brandsxai_leads.insert_one({
        "id": new_id, "name": lead.name, "email": lead.email, "company": lead.company,
        "message": lead.message, "created_at": datetime.now(timezone.utc)
    })
    return {"message": "Lead created", "id": new_id, "db_source": "mongodb"}

# ==================== CAMPAIGNS ====================

class CampaignCreate(BaseModel):
    name: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    target_audience: Optional[str] = None
    call_script: Optional[str] = None

class OpportunityCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    business_name: Optional[str] = None
    notes: Optional[str] = None

class OpportunityStageUpdate(BaseModel):
    stage: str

class OpportunityUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    business_name: Optional[str] = None
    notes: Optional[str] = None
    call_summary: Optional[str] = None
    recording_url: Optional[str] = None
    call_duration: Optional[int] = None
    call_outcome: Optional[str] = None

@api_router.get("/campaigns")
async def get_campaigns(current_user: dict = Depends(get_current_user)):
    """Get all campaigns for the user's brand"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    if not brand_id:
        raise HTTPException(status_code=400, detail="User has no brand assigned")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT c.*, 
                           COUNT(o.id) as total_opportunities,
                           COALESCE(SUM(o.opportunity_value), 0) as total_value
                    FROM brandsxai_campaigns c
                    LEFT JOIN brandsxai_opportunities o ON c.id = o.campaign_id
                    WHERE c.brand_id = %s
                    GROUP BY c.id
                    ORDER BY c.created_at DESC
                """, (brand_id,))
                campaigns = cursor.fetchall()
                return {"campaigns": campaigns, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get campaigns error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    campaigns = await mongo_db.brandsxai_campaigns.find(
        {"brand_id": brand_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Add opportunity stats
    for c in campaigns:
        opps = await mongo_db.brandsxai_opportunities.find({"campaign_id": c['id']}, {"_id": 0}).to_list(1000)
        c['total_opportunities'] = len(opps)
        c['total_value'] = sum(o.get('opportunity_value', 0) for o in opps)
    
    return {"campaigns": campaigns, "db_source": "mongodb"}

@api_router.post("/campaigns")
async def create_campaign(campaign: CampaignCreate, current_user: dict = Depends(get_current_user)):
    """Create a new campaign"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    user_id = int(current_user.get('sub'))
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO brandsxai_campaigns 
                    (brand_id, name, description, start_date, end_date, target_audience, call_script, created_by, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'active')
                """, (brand_id, campaign.name, campaign.description, campaign.start_date, 
                      campaign.end_date, campaign.target_audience, campaign.call_script, user_id))
                mysql_conn.commit()
                campaign_id = cursor.lastrowid
                
                cursor.execute("SELECT * FROM brandsxai_campaigns WHERE id = %s", (campaign_id,))
                result = cursor.fetchone()
                return {"campaign": result, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL create campaign error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    last_campaign = await mongo_db.brandsxai_campaigns.find_one(sort=[("id", -1)])
    new_id = (last_campaign.get('id', 0) + 1) if last_campaign else 1
    
    new_campaign = {
        "id": new_id,
        "brand_id": brand_id,
        "name": campaign.name,
        "description": campaign.description,
        "start_date": campaign.start_date,
        "end_date": campaign.end_date,
        "target_audience": campaign.target_audience,
        "call_script": campaign.call_script,
        "status": "active",
        "created_by": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await mongo_db.brandsxai_campaigns.insert_one(new_campaign)
    new_campaign.pop('_id', None)
    
    return {"campaign": new_campaign, "db_source": "mongodb"}

@api_router.get("/campaigns/{campaign_id}")
async def get_campaign(campaign_id: int, current_user: dict = Depends(get_current_user)):
    """Get campaign details with opportunities"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_campaigns WHERE id = %s AND brand_id = %s", (campaign_id, brand_id))
                campaign = cursor.fetchone()
                if not campaign:
                    raise HTTPException(status_code=404, detail="Campaign not found")
                
                cursor.execute("SELECT * FROM brandsxai_opportunities WHERE campaign_id = %s ORDER BY created_at DESC", (campaign_id,))
                opportunities = cursor.fetchall()
                
                # Group by stage
                stages = {}
                for stage in ['dialing', 'interested', 'not_interested', 'callback', 'store_visit', 'invalid_number']:
                    stage_opps = [o for o in opportunities if o['stage'] == stage]
                    stages[stage] = {
                        'opportunities': stage_opps,
                        'count': len(stage_opps),
                        'total_value': sum(float(o['opportunity_value'] or 0) for o in stage_opps)
                    }
                
                return {"campaign": campaign, "stages": stages, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL get campaign error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    campaign = await mongo_db.brandsxai_campaigns.find_one({"id": campaign_id, "brand_id": brand_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    opportunities = await mongo_db.brandsxai_opportunities.find({"campaign_id": campaign_id}, {"_id": 0}).to_list(1000)
    
    stages = {}
    for stage in ['dialing', 'interested', 'not_interested', 'callback', 'store_visit', 'invalid_number']:
        stage_opps = [o for o in opportunities if o.get('stage') == stage]
        stages[stage] = {
            'opportunities': stage_opps,
            'count': len(stage_opps),
            'total_value': sum(o.get('opportunity_value', 0) for o in stage_opps)
        }
    
    return {"campaign": campaign, "stages": stages, "db_source": "mongodb"}

@api_router.post("/campaigns/{campaign_id}/opportunities")
async def create_opportunity(campaign_id: int, opportunity: OpportunityCreate, current_user: dict = Depends(get_current_user)):
    """Add an opportunity to a campaign"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                # Verify campaign belongs to user's brand
                cursor.execute("SELECT id FROM brandsxai_campaigns WHERE id = %s AND brand_id = %s", (campaign_id, brand_id))
                if not cursor.fetchone():
                    raise HTTPException(status_code=404, detail="Campaign not found")
                
                cursor.execute("""
                    INSERT INTO brandsxai_opportunities 
                    (campaign_id, brand_id, name, phone, email, business_name, notes, stage)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'dialing')
                """, (campaign_id, brand_id, opportunity.name, opportunity.phone, opportunity.email,
                      opportunity.business_name, opportunity.notes))
                mysql_conn.commit()
                
                opp_id = cursor.lastrowid
                cursor.execute("SELECT * FROM brandsxai_opportunities WHERE id = %s", (opp_id,))
                result = cursor.fetchone()
                return {"opportunity": result, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL create opportunity error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    campaign = await mongo_db.brandsxai_campaigns.find_one({"id": campaign_id, "brand_id": brand_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    last_opp = await mongo_db.brandsxai_opportunities.find_one(sort=[("id", -1)])
    new_id = (last_opp.get('id', 0) + 1) if last_opp else 1
    
    new_opp = {
        "id": new_id,
        "campaign_id": campaign_id,
        "brand_id": brand_id,
        "name": opportunity.name,
        "phone": opportunity.phone,
        "email": opportunity.email,
        "business_name": opportunity.business_name,
        "notes": opportunity.notes,
        "stage": "dialing",
        "call_summary": None,
        "recording_url": None,
        "call_duration": None,
        "call_outcome": None,
        "last_called_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await mongo_db.brandsxai_opportunities.insert_one(new_opp)
    new_opp.pop('_id', None)
    
    return {"opportunity": new_opp, "db_source": "mongodb"}

@api_router.put("/opportunities/{opportunity_id}/stage")
async def update_opportunity_stage(opportunity_id: int, update: OpportunityStageUpdate, current_user: dict = Depends(get_current_user)):
    """Update opportunity stage (for drag-drop in Kanban)"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    valid_stages = ['dialing', 'interested', 'not_interested', 'callback', 'store_visit', 'invalid_number']
    
    if update.stage not in valid_stages:
        raise HTTPException(status_code=400, detail=f"Invalid stage. Must be one of: {valid_stages}")
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute(
                    "UPDATE brandsxai_opportunities SET stage = %s, updated_at = NOW() WHERE id = %s AND brand_id = %s",
                    (update.stage, opportunity_id, brand_id)
                )
                mysql_conn.commit()
                
                if cursor.rowcount == 0:
                    raise HTTPException(status_code=404, detail="Opportunity not found")
                
                cursor.execute("SELECT * FROM brandsxai_opportunities WHERE id = %s", (opportunity_id,))
                result = cursor.fetchone()
                return {"opportunity": result, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL update stage error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    result = await mongo_db.brandsxai_opportunities.update_one(
        {"id": opportunity_id, "brand_id": brand_id},
        {"$set": {"stage": update.stage, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Opportunity not found")
    
    opp = await mongo_db.brandsxai_opportunities.find_one({"id": opportunity_id}, {"_id": 0})
    return {"opportunity": opp, "db_source": "mongodb"}

@api_router.get("/opportunities/{opportunity_id}")
async def get_opportunity(opportunity_id: int, current_user: dict = Depends(get_current_user)):
    """Get single opportunity details"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_opportunities WHERE id = %s AND brand_id = %s", (opportunity_id, brand_id))
                result = cursor.fetchone()
                if not result:
                    raise HTTPException(status_code=404, detail="Opportunity not found")
                return {"opportunity": result, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL get opportunity error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    opp = await mongo_db.brandsxai_opportunities.find_one({"id": opportunity_id, "brand_id": brand_id}, {"_id": 0})
    if not opp:
        raise HTTPException(status_code=404, detail="Opportunity not found")
    return {"opportunity": opp, "db_source": "mongodb"}

@api_router.put("/opportunities/{opportunity_id}")
async def update_opportunity(opportunity_id: int, update: OpportunityUpdate, current_user: dict = Depends(get_current_user)):
    """Update opportunity details"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    # Build update dict with non-None values
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                set_clause = ", ".join([f"{k} = %s" for k in update_data.keys()])
                values = list(update_data.values()) + [opportunity_id, brand_id]
                cursor.execute(
                    f"UPDATE brandsxai_opportunities SET {set_clause} WHERE id = %s AND brand_id = %s",
                    values
                )
                mysql_conn.commit()
                
                if cursor.rowcount == 0:
                    raise HTTPException(status_code=404, detail="Opportunity not found")
                
                cursor.execute("SELECT * FROM brandsxai_opportunities WHERE id = %s", (opportunity_id,))
                result = cursor.fetchone()
                return {"opportunity": result, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL update opportunity error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    result = await mongo_db.brandsxai_opportunities.update_one(
        {"id": opportunity_id, "brand_id": brand_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        # Check if exists
        existing = await mongo_db.brandsxai_opportunities.find_one({"id": opportunity_id, "brand_id": brand_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Opportunity not found")
    
    opp = await mongo_db.brandsxai_opportunities.find_one({"id": opportunity_id}, {"_id": 0})
    return {"opportunity": opp, "db_source": "mongodb"}

class DialRequest(BaseModel):
    phone: str

# AI Voice Calling API Configuration
AI_CALL_API_URL = "http://16.16.213.64:8000/api/call"
AI_CALL_API_KEY = "kj-neha-2024-xK9p"

def validate_phone_number(phone: str) -> tuple[bool, str]:
    """Validate phone number format - must be 10 digits (excluding country code)"""
    import re
    # Remove all non-digit characters except + at the start
    cleaned = re.sub(r'[^\d+]', '', phone)
    
    # Check if it has country code (+91 or 91) and 10 digit number
    if cleaned.startswith('+91'):
        digits = cleaned[3:]  # Remove +91
    elif cleaned.startswith('91') and len(cleaned) >= 12:
        digits = cleaned[2:]  # Remove 91
    else:
        digits = cleaned.lstrip('+')  # Just the digits
    
    # Check if we have exactly 10 digits
    if len(digits) == 10 and digits.isdigit():
        # Format properly with country code
        formatted = f"+91{digits}"
        return True, formatted
    
    return False, "Phone number must be 10 digits (excluding country code)"

@api_router.post("/opportunities/{opportunity_id}/dial")
async def dial_opportunity(opportunity_id: int, dial_req: DialRequest, current_user: dict = Depends(get_current_user)):
    """Initiate AI voice call to an opportunity using external AI calling service"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    # Validate phone number
    is_valid, result = validate_phone_number(dial_req.phone)
    if not is_valid:
        raise HTTPException(status_code=400, detail=result)
    
    formatted_phone = result
    
    # Call the AI Voice Calling API
    import httpx
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                AI_CALL_API_URL,
                headers={
                    "Content-Type": "application/json",
                    "X-API-Key": AI_CALL_API_KEY
                },
                json={"phone_number": formatted_phone}
            )
            
            if response.status_code == 200:
                call_response = response.json()
                
                if call_response.get('success'):
                    # Call initiated successfully
                    update_data = {
                        'call_id': call_response.get('call_id'),
                        'last_called_at': datetime.now(timezone.utc).isoformat(),
                        'updated_at': datetime.now(timezone.utc).isoformat()
                    }
                    
                    # Update MongoDB
                    await mongo_db.brandsxai_opportunities.update_one(
                        {"id": opportunity_id, "brand_id": brand_id},
                        {"$set": update_data}
                    )
                    
                    return {
                        "success": True,
                        "message": "Call initiated successfully",
                        "call_id": call_response.get('call_id'),
                        "room_name": call_response.get('room_name'),
                        "phone_number": formatted_phone,
                        "timestamp": call_response.get('timestamp')
                    }
                else:
                    raise HTTPException(status_code=400, detail="AI service returned failure")
            else:
                error_detail = f"AI calling service error: HTTP {response.status_code}"
                try:
                    error_body = response.json()
                    if 'detail' in error_body:
                        error_detail = error_body['detail']
                except:
                    pass
                raise HTTPException(status_code=502, detail=error_detail)
                
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="AI calling service timeout - please try again")
    except httpx.RequestError as e:
        logger.error(f"AI call API request error: {e}")
        raise HTTPException(status_code=502, detail="Failed to connect to AI calling service")

@api_router.get("/contacts")
async def get_all_contacts(current_user: dict = Depends(get_current_user)):
    """Get all contacts/opportunities across all campaigns for the user's brand"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT o.*, c.name as campaign_name 
                    FROM brandsxai_opportunities o
                    JOIN brandsxai_campaigns c ON o.campaign_id = c.id
                    WHERE o.brand_id = %s
                    ORDER BY o.created_at DESC
                """, (brand_id,))
                contacts = cursor.fetchall()
                return {"contacts": contacts, "count": len(contacts), "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get contacts error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    # First get campaign names
    campaigns = await mongo_db.brandsxai_campaigns.find({"brand_id": brand_id}, {"_id": 0}).to_list(100)
    campaign_map = {c['id']: c['name'] for c in campaigns}
    
    # Get all opportunities
    opportunities = await mongo_db.brandsxai_opportunities.find(
        {"brand_id": brand_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    # Add campaign names
    for opp in opportunities:
        opp['campaign_name'] = campaign_map.get(opp.get('campaign_id'), 'Unknown')
    
    return {"contacts": opportunities, "count": len(opportunities), "db_source": "mongodb"}

# ==================== SESSION/CALLS ====================

@api_router.get("/sessions/calls")
async def get_session_calls(current_user: dict = Depends(get_current_user)):
    """Get all call sessions for the user's brand"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT * FROM brandsxai_calls 
                    WHERE brand_id = %s 
                    ORDER BY started_at DESC 
                    LIMIT 100
                """, (brand_id,))
                calls = cursor.fetchall()
                
                total_duration = sum(c.get('duration', 0) or 0 for c in calls)
                avg_duration = total_duration // len(calls) if calls else 0
                active_calls = len([c for c in calls if c.get('status') == 'active'])
                calls_with_issues = len([c for c in calls if c.get('has_issue')])
                
                return {
                    "calls": calls,
                    "stats": {
                        "totalCalls": len(calls),
                        "totalDuration": total_duration,
                        "avgDuration": avg_duration,
                        "activeCalls": active_calls,
                        "callsWithIssues": calls_with_issues
                    },
                    "db_source": "mysql"
                }
        except Exception as e:
            logger.error(f"MySQL get calls error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    calls = await mongo_db.brandsxai_calls.find(
        {"brand_id": brand_id}, {"_id": 0}
    ).sort("started_at", -1).limit(100).to_list(100)
    
    total_duration = sum(c.get('duration', 0) or 0 for c in calls)
    avg_duration = total_duration // len(calls) if calls else 0
    active_calls = len([c for c in calls if c.get('status') == 'active'])
    calls_with_issues = len([c for c in calls if c.get('has_issue')])
    
    return {
        "calls": calls,
        "stats": {
            "totalCalls": len(calls),
            "totalDuration": total_duration,
            "avgDuration": avg_duration,
            "activeCalls": active_calls,
            "callsWithIssues": calls_with_issues
        },
        "db_source": "mongodb"
    }

# ==================== CLAIM PROCESSING ====================

# ICD-10 Code validation data (common codes for autocomplete)
ICD10_COMMON_CODES = {
    "E11.9": "Type 2 diabetes mellitus without complications",
    "I10": "Essential (primary) hypertension",
    "J06.9": "Acute upper respiratory infection, unspecified",
    "K21.0": "Gastro-esophageal reflux disease with esophagitis",
    "K21.9": "Gastro-esophageal reflux disease without esophagitis",
    "E53.8": "Deficiency of other specified B group vitamins",
    "Z98.84": "Bariatric surgery status",
    "R05.9": "Cough, unspecified",
    "M54.5": "Low back pain",
    "F32.9": "Major depressive disorder, single episode, unspecified",
    "J45.909": "Unspecified asthma, uncomplicated",
    "E78.5": "Hyperlipidemia, unspecified",
    "N39.0": "Urinary tract infection, site not specified",
    "J18.9": "Pneumonia, unspecified organism",
    "G43.909": "Migraine, unspecified, not intractable, without status migrainosus",
    "K59.00": "Constipation, unspecified",
    "R10.9": "Unspecified abdominal pain",
    "R51.9": "Headache, unspecified",
    "Z87.891": "Personal history of nicotine dependence",
    "Z79.899": "Other long term (current) drug therapy",
}

ICD10_EXTRACTION_PROMPT = """You are an expert medical coder specializing in ICD-10 code extraction from clinical notes.

Your task is to analyze the provided clinical document and extract ALL relevant ICD-10 codes.

IMPORTANT GUIDELINES:
1. Extract codes for ALL documented conditions, diagnoses, and symptoms
2. Include codes from ALL sections: Chief Complaint, History, Assessment & Plan, Review of Systems, etc.
3. Include both primary and secondary diagnoses
4. Include codes for chronic conditions mentioned (even if "controlled" or "stable")
5. Include codes for surgical history (Z codes)
6. Include codes for current medications context
7. Be thorough - it's better to include a code that can be removed than to miss one

For each code, provide:
- code: The ICD-10 code (e.g., "E11.9")
- description: Brief description of what the code represents
- source_text: The exact phrase or sentence from the document that supports this code
- confidence: Your confidence level (0.0-1.0)

Respond ONLY with a JSON array of codes. No other text.
Example format:
[
  {"code": "E11.9", "description": "Type 2 diabetes mellitus without complications", "source_text": "Patient has history of DM2", "confidence": 0.95},
  {"code": "I10", "description": "Essential hypertension", "source_text": "HTN controlled on lisinopril", "confidence": 0.9}
]

If no valid codes can be extracted, return an empty array: []
"""

@api_router.post("/claim-processing/sessions")
async def create_claim_session(
    request: ClaimSessionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new claim processing session"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    import uuid
    session_id = str(uuid.uuid4())
    user_id = int(current_user.get('sub'))
    brand_id = current_user.get('brand_id')
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO brandsxai_claim_sessions (id, user_id, brand_id, title, extracted_codes)
                    VALUES (%s, %s, %s, %s, %s)
                """, (session_id, user_id, brand_id, request.title, json.dumps([])))
                mysql_conn.commit()
                
                return {
                    "id": session_id,
                    "title": request.title,
                    "status": "active",
                    "extracted_codes": [],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "db_source": "mysql"
                }
        except Exception as e:
            logger.error(f"MySQL create claim session error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    session = {
        "id": session_id,
        "user_id": user_id,
        "brand_id": brand_id,
        "title": request.title,
        "status": "active",
        "extracted_codes": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await mongo_db.brandsxai_claim_sessions.insert_one(session)
    del session['_id']
    session['db_source'] = 'mongodb'
    return session

@api_router.get("/claim-processing/sessions")
async def get_claim_sessions(current_user: dict = Depends(get_current_user)):
    """Get all claim sessions for the user"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    user_id = int(current_user.get('sub'))
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT id, title, status, extracted_codes, created_at, updated_at
                    FROM brandsxai_claim_sessions 
                    WHERE user_id = %s 
                    ORDER BY updated_at DESC
                    LIMIT 50
                """, (user_id,))
                sessions = cursor.fetchall()
                
                for s in sessions:
                    if isinstance(s.get('extracted_codes'), str):
                        s['extracted_codes'] = json.loads(s['extracted_codes'])
                    s['created_at'] = s['created_at'].isoformat() if s.get('created_at') else None
                    s['updated_at'] = s['updated_at'].isoformat() if s.get('updated_at') else None
                
                return {"sessions": sessions, "db_source": "mysql"}
        except Exception as e:
            logger.error(f"MySQL get claim sessions error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    sessions = await mongo_db.brandsxai_claim_sessions.find(
        {"user_id": user_id}, {"_id": 0}
    ).sort("updated_at", -1).limit(50).to_list(50)
    
    return {"sessions": sessions, "db_source": "mongodb"}

@api_router.get("/claim-processing/sessions/{session_id}")
async def get_claim_session(session_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific claim session with messages"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    user_id = int(current_user.get('sub'))
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT * FROM brandsxai_claim_sessions 
                    WHERE id = %s AND user_id = %s
                """, (session_id, user_id))
                session = cursor.fetchone()
                
                if not session:
                    raise HTTPException(status_code=404, detail="Session not found")
                
                cursor.execute("""
                    SELECT id, role, content, file_info, codes_extracted, created_at
                    FROM brandsxai_claim_messages 
                    WHERE session_id = %s 
                    ORDER BY created_at ASC
                """, (session_id,))
                messages = cursor.fetchall()
                
                if isinstance(session.get('extracted_codes'), str):
                    session['extracted_codes'] = json.loads(session['extracted_codes'])
                session['created_at'] = session['created_at'].isoformat() if session.get('created_at') else None
                session['updated_at'] = session['updated_at'].isoformat() if session.get('updated_at') else None
                
                for m in messages:
                    if isinstance(m.get('file_info'), str):
                        m['file_info'] = json.loads(m['file_info'])
                    if isinstance(m.get('codes_extracted'), str):
                        m['codes_extracted'] = json.loads(m['codes_extracted'])
                    m['created_at'] = m['created_at'].isoformat() if m.get('created_at') else None
                
                return {"session": session, "messages": messages, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL get claim session error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    session = await mongo_db.brandsxai_claim_sessions.find_one(
        {"id": session_id, "user_id": user_id}, {"_id": 0}
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    messages = await mongo_db.brandsxai_claim_messages.find(
        {"session_id": session_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(1000)
    
    return {"session": session, "messages": messages, "db_source": "mongodb"}

@api_router.post("/claim-processing/sessions/{session_id}/chat")
async def chat_with_claim_session(
    session_id: str,
    request: ClaimMessage,
    current_user: dict = Depends(get_current_user)
):
    """Send a message to the claim processing AI"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    user_id = int(current_user.get('sub'))
    
    # Verify session exists
    session = None
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_claim_sessions WHERE id = %s AND user_id = %s", (session_id, user_id))
                session = cursor.fetchone()
        except Exception as e:
            logger.error(f"MySQL session check error: {e}")
        finally:
            mysql_conn.close()
    
    if not session:
        session = await mongo_db.brandsxai_claim_sessions.find_one({"id": session_id, "user_id": user_id})
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get existing codes from session
    existing_codes = session.get('extracted_codes', [])
    if isinstance(existing_codes, str):
        existing_codes = json.loads(existing_codes)
    
    # Prepare the AI prompt
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContent
        import uuid as uuid_module
        
        emergent_key = os.environ.get('EMERGENT_LLM_KEY')
        if not emergent_key:
            raise HTTPException(status_code=500, detail="AI service not configured")
        
        # Build system message for the AI
        system_message = """You are an expert ICD-10 medical coding assistant. 
Your task is to help extract and manage ICD-10 codes from clinical documents.
When users provide documents, extract all relevant ICD-10 codes.
When users ask questions or provide feedback, respond helpfully."""
        
        # Initialize chat with proper method chaining
        chat = LlmChat(
            api_key=emergent_key,
            session_id=f"claim-{session_id}-{uuid_module.uuid4().hex[:8]}",
            system_message=system_message
        ).with_model("gemini", "gemini-2.5-flash")
        
        # Build context message
        context = f"""Current session has these extracted codes: {json.dumps(existing_codes)}

User message: {request.content}

{ICD10_EXTRACTION_PROMPT if request.file_data else ""}

If the user is asking a question, answer it helpfully. 
If they're providing feedback about codes (wrong, missing, etc.), acknowledge and provide the corrected code(s).
If they attach documents, extract all ICD-10 codes.

Always respond in this JSON format:
{{
  "response_text": "Your helpful response to the user",
  "new_codes": [
    {{"code": "X00.0", "description": "...", "source_text": "...", "confidence": 0.9}}
  ],
  "codes_to_remove": ["X00.0"]
}}
"""
        
        # Build message with file content if present
        file_contents = []
        file_info = []
        if request.file_data:
            for f in request.file_data:
                file_contents.append(FileContent(
                    content_type=f.get('content_type', 'application/pdf'),
                    file_content_base64=f.get('base64_data', '')
                ))
                file_info.append({
                    "filename": f.get('filename', 'document'),
                    "content_type": f.get('content_type')
                })
        
        user_msg = UserMessage(text=context, file_contents=file_contents if file_contents else None)
        
        # Send to AI
        response = await chat.send_message(user_msg)
        response_text = str(response)
        
        # Parse AI response
        try:
            # Try to extract JSON from response
            import re
            json_match = re.search(r'\{[\s\S]*\}', response_text)
            if json_match:
                ai_response = json.loads(json_match.group())
            else:
                ai_response = {"response_text": response_text, "new_codes": [], "codes_to_remove": []}
        except json.JSONDecodeError:
            ai_response = {"response_text": response_text, "new_codes": [], "codes_to_remove": []}
        
        assistant_text = ai_response.get('response_text', response_text)
        new_codes = ai_response.get('new_codes', [])
        codes_to_remove = ai_response.get('codes_to_remove', [])
        
        # Update codes list
        updated_codes = [c for c in existing_codes if c.get('code') not in codes_to_remove]
        existing_code_set = {c.get('code') for c in updated_codes}
        for nc in new_codes:
            if nc.get('code') and nc.get('code') not in existing_code_set:
                updated_codes.append(nc)
        
        # Save to database
        mysql_conn = try_mysql_connection()
        if mysql_conn:
            try:
                with mysql_conn.cursor() as cursor:
                    # Save user message
                    cursor.execute("""
                        INSERT INTO brandsxai_claim_messages (session_id, role, content, file_info)
                        VALUES (%s, 'user', %s, %s)
                    """, (session_id, request.content, json.dumps(file_info) if file_info else None))
                    
                    # Save assistant message
                    cursor.execute("""
                        INSERT INTO brandsxai_claim_messages (session_id, role, content, codes_extracted)
                        VALUES (%s, 'assistant', %s, %s)
                    """, (session_id, assistant_text, json.dumps(new_codes) if new_codes else None))
                    
                    # Update session codes
                    cursor.execute("""
                        UPDATE brandsxai_claim_sessions 
                        SET extracted_codes = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (json.dumps(updated_codes), session_id))
                    
                    mysql_conn.commit()
            except Exception as e:
                logger.error(f"MySQL save chat error: {e}")
            finally:
                mysql_conn.close()
        else:
            # MongoDB fallback
            await mongo_db.brandsxai_claim_messages.insert_one({
                "session_id": session_id,
                "role": "user",
                "content": request.content,
                "file_info": file_info if file_info else None,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            await mongo_db.brandsxai_claim_messages.insert_one({
                "session_id": session_id,
                "role": "assistant",
                "content": assistant_text,
                "codes_extracted": new_codes if new_codes else None,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            await mongo_db.brandsxai_claim_sessions.update_one(
                {"id": session_id},
                {"$set": {"extracted_codes": updated_codes, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        
        return {
            "response": assistant_text,
            "new_codes": new_codes,
            "codes_removed": codes_to_remove,
            "all_codes": updated_codes
        }
        
    except ImportError as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail="AI service not available")
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing request: {str(e)}")

@api_router.put("/claim-processing/sessions/{session_id}/codes")
async def update_session_codes(
    session_id: str,
    request: CodeUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Manually update the codes in a session"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    user_id = int(current_user.get('sub'))
    
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("""
                    UPDATE brandsxai_claim_sessions 
                    SET extracted_codes = %s, updated_at = NOW()
                    WHERE id = %s AND user_id = %s
                """, (json.dumps(request.codes), session_id, user_id))
                mysql_conn.commit()
                
                if cursor.rowcount == 0:
                    raise HTTPException(status_code=404, detail="Session not found")
                
                return {"success": True, "codes": request.codes, "db_source": "mysql"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"MySQL update codes error: {e}")
        finally:
            mysql_conn.close()
    
    # MongoDB fallback
    result = await mongo_db.brandsxai_claim_sessions.update_one(
        {"id": session_id, "user_id": user_id},
        {"$set": {"extracted_codes": request.codes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"success": True, "codes": request.codes, "db_source": "mongodb"}

@api_router.get("/claim-processing/icd10/search")
async def search_icd10_codes(q: str = "", current_user: dict = Depends(get_current_user)):
    """Search ICD-10 codes for autocomplete"""
    if not current_user:
        raise HTTPException(status_code=403, detail="Authentication required")
    
    if len(q) < 2:
        return {"codes": []}
    
    q_lower = q.lower()
    results = []
    
    for code, desc in ICD10_COMMON_CODES.items():
        if q_lower in code.lower() or q_lower in desc.lower():
            results.append({"code": code, "description": desc})
    
    return {"codes": results[:20]}

@api_router.get("/claim-processing/sessions/{session_id}/export")
async def export_session_codes(session_id: str, current_user: dict = Depends(get_current_user)):
    """Export session codes as Excel file"""
    if not current_user or current_user.get('type') == 'admin':
        raise HTTPException(status_code=403, detail="User access required")
    
    user_id = int(current_user.get('sub'))
    
    # Get session
    session = None
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        try:
            with mysql_conn.cursor() as cursor:
                cursor.execute("SELECT * FROM brandsxai_claim_sessions WHERE id = %s AND user_id = %s", (session_id, user_id))
                session = cursor.fetchone()
        except Exception as e:
            logger.error(f"MySQL export error: {e}")
        finally:
            mysql_conn.close()
    
    if not session:
        session = await mongo_db.brandsxai_claim_sessions.find_one({"id": session_id, "user_id": user_id})
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    codes = session.get('extracted_codes', [])
    if isinstance(codes, str):
        codes = json.loads(codes)
    
    # Create Excel file
    import io
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ICD-10 Codes"
        
        # Header row
        ws['A1'] = "ICD-10 Codes"
        
        # Single row with all codes
        code_list = [c.get('code', '') for c in codes]
        ws['A2'] = ", ".join(code_list)
        
        # Second sheet with details
        ws2 = wb.create_sheet("Code Details")
        ws2['A1'] = "Code"
        ws2['B1'] = "Description"
        ws2['C1'] = "Source Text"
        
        for i, code in enumerate(codes, start=2):
            ws2[f'A{i}'] = code.get('code', '')
            ws2[f'B{i}'] = code.get('description', '')
            ws2[f'C{i}'] = code.get('source_text', '')
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=icd10_codes_{session_id[:8]}.xlsx"}
        )
    except ImportError:
        # Fallback to CSV
        import csv
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["ICD-10 Codes"])
        writer.writerow([", ".join([c.get('code', '') for c in codes])])
        
        from fastapi.responses import Response
        return Response(
            content=buffer.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=icd10_codes_{session_id[:8]}.csv"}
        )

# ==================== UTILITY ====================

@api_router.get("/")
async def root():
    return {"message": "BrandsXAI API", "version": "2.0.0"}

@api_router.get("/db-status")
async def db_status():
    mysql_status = try_mysql_connection() is not None
    return {"mysql_available": mysql_status, "mongodb_available": True, "primary_db": "mysql" if mysql_status else "mongodb"}

# Include router & middleware
app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
async def startup_event():
    mysql_conn = try_mysql_connection()
    if mysql_conn:
        init_mysql_tables(mysql_conn)
        mysql_conn.close()
    await init_mongodb_collections()
    logger.info("Database initialization complete")
